import select

import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import time
from selenium.webdriver.support.ui import Select
import re
from datetime import date

# 為了打開Chrome
option = webdriver.ChromeOptions()
option.add_experimental_option("detach", True)
chrome = r'C://Program Files (x86)//Google//Chrome//Application//chromedriver.exe' # 本機chromedriver的位置
driver = webdriver.Chrome(executable_path=chrome, options=option)

# browser
# driver.set_window_size(802.6, 721.4)

# today info
today = date.today().strftime('%Y%m%d')
today = today.lstrip('-')
print(today)

all_person = []
def open_name_list():
    # name.xlsx should provide ['idCardNum', 'monthOfBirth', 'DayOfBirth']
    # no need header in name.xlsx
    wb = openpyxl.load_workbook('name.xlsx')
    sheet = wb.worksheets[0]
    num_rows = sheet.max_row
    num_columns = sheet.max_column
    print(f'num of rows: {num_rows}')
    print((f'num of columns: {num_columns}'))

    # get data
    for i in range(1, num_rows+1):
        this_person = []
        for j in range(1, num_columns+1):
            this_person.append(sheet.cell(row=i, column=j).value)
        all_person.append(this_person)
    #print(all_person)

graduated = []
non_graduated = []

def check_graduated(person_info):
    # 前往網頁
    driver.get('https://wapps2.ipm.edu.mo/rvdweb/pd_student_login.asp')

    # personal information
    idcard = person_info[0]
    birth_month = person_info[1]
    birth_day = person_info[2]

    driver.find_element_by_xpath('/html/body/form/table[1]/tbody/tr[3]/td[2]/input').send_keys(idcard)

    select_month = Select(driver.find_element_by_xpath('/html/body/form/table[1]/tbody/tr[6]/td[2]/select'))
    select_month.select_by_value(birth_month)

    select_day = Select(driver.find_element_by_xpath('/html/body/form/table[1]/tbody/tr[6]/td[2]/select[2]'))
    select_day.select_by_value(birth_day)

    driver.find_element_by_xpath('/html/body/form/table[1]/tbody/tr[8]/td[2]/input').click()

    # get id
    stu_id = driver.find_element_by_xpath('/html/body/div[2]/form/table[1]/tbody/tr[1]/td[2]').get_attribute(
        "innerHTML")
    stu_id = stu_id.strip()  # remove spacing
    stu_id = stu_id[:11]  # 因為html有太多不可行
    stu_id = stu_id.replace('-', '')
    # print(stu_id)

    # get name
    stu_name = driver.find_element_by_xpath('/html/body/div[2]/form/table[1]/tbody/tr[2]/td[2]').get_attribute(
        "innerHTML")
    stu_name = stu_name.strip()  # remove spacing

    # get programme
    progs_select = driver.find_element_by_xpath('/html/body/div[2]/form/table[1]/tbody/tr[3]/td[2]/font/select')
    prog_lists = [x for x in progs_select.find_elements_by_tag_name('option')]
    prog_lists = reversed(prog_lists)

    # programme state, 因為新舊生版本有出入, 所以舊生要在外部取值
    prog_state = driver.find_element_by_xpath(
        '/html/body/div[2]/form/table[1]/tbody/tr[3]/td[2]/font/input[4]').get_attribute('value')


    my_list = []

    for p in prog_lists:
        # 由古老記錄開始找
        prog_name = p.get_attribute('text')
        prog_name = prog_name.split('(')
        prog_name = prog_name[0].strip()
        prog_name = prog_name.lower().title()

        prog_val = p.get_attribute('value')
        prog_val = prog_val.replace('$', '#')
        prog_val_str = prog_val.split('#')

        # 因為發現舊生的code不同
        if len(prog_val_str) > 1:
            prog_code = prog_val.split('#')[0]
            prog_state = prog_val.split('#')[1]
            faculty = prog_val.split('#')[2].upper()
        else:
            prog_code = prog_val.split('#')[0]
            faculty = None  # 舊生沒有school/faculty的值, 在讀生都不一定有

        # print(f'Programme code: {prog_code}')
        # print(f'Programme state: {prog_state}')
        # print(f'Faculty: {faculty}')

        my_list.append([stu_id, stu_name, faculty, prog_code, prog_name, prog_state])

    # 反轉my_list, 找出最近的一次畢業
    my_list = my_list[::-1]

    for i in range(0, len(my_list)):
        if my_list[i][-1] == 'C':
            graduated.append(my_list[i])
            break
        if i == len(my_list) - 1:
            print(f'This person does not graduated yet.')
            non_graduated.append(my_list[i])

def save_graduated(graduated, non_graduated):
    headers = ['Student ID', 'Name', 'Faculty/School', 'Programme Code', 'Programme', 'State']
    excel_name = f'checked_{today}.xlsx'
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = 'graduated'
    ws1.append(headers)
    for i in graduated:
        ws1.append(i)
    ws2 = wb.create_sheet('non_graduated')
    ws2.append(headers)
    for i in non_graduated:
        ws2.append(i)
    wb.save(excel_name)


# run
open_name_list()
for i in range(0, len(all_person)):
    person_info = all_person[i]
    check_graduated(person_info)

print(f'graduated: {graduated}')
print(f'non_graduated: {non_graduated}')

save_graduated(graduated, non_graduated)

driver.close()
